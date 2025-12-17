# Импорт необходимых библиотек и модулей
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from fpdf import FPDF
from openpyxl import Workbook
from PIL import Image, ImageTk
import os
import webbrowser
import urllib.parse


# Определение класса Text_Editor
class Text_Editor:
    # Определение конструктора
    def __init__(self, root):
        # Назначение root
        self.root = root
        # Заголовок окна
        self.root.title("Блокнот")
        # Геометрия окна
        self.root.geometry("640x480+150+100")
        # Инициализация имени файла
        self.filename = None
        # Объявление переменной Title
        self.title = StringVar()
        # Объявление переменной состояния
        self.status = StringVar()

        # Создание заголовка
        self.titlebar = Label(
            self.root,
            textvariable=self.title,
            font=("times new roman", 15, "bold"),
            bd=2,
            relief=GROOVE
        )
        self.titlebar.pack(side=TOP, fill=BOTH)

        # Вызов функции settitle
        self.settitle()

        # Создание строки состояния
        self.statusbar = Label(
            self.root,
            textvariable=self.status,
            font=("times new roman", 15, "bold"),
            bd=2,
            relief=GROOVE
        )
        self.statusbar.pack(side=BOTTOM, fill=BOTH)
        self.status.set("Добро пожаловать в текстовый редактор")

        # Создание панели меню
        self.menubar = Menu(self.root, font=("times new roman", 15, "bold"), activebackground="skyblue")
        self.root.config(menu=self.menubar)

        # Меню "Файл"
        self.filemenu = Menu(self.menubar, font=("times new roman", 12, "bold"), activebackground="skyblue", tearoff=0)
        self.filemenu.add_command(label="Новый", accelerator="Ctrl+N", command=self.newfile)
        self.filemenu.add_command(label="Открыть", accelerator="Ctrl+O", command=self.openfile)
        self.filemenu.add_command(label="Сохранить", accelerator="Ctrl+S", command=self.savefile)
        self.filemenu.add_command(label="Сохранить как", accelerator="Ctrl+A", command=self.saveasfile)
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Сохранить как PDF", command=self.save_as_pdf)
        self.filemenu.add_command(label="Сохранить как Excel", command=self.save_as_excel)
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Закрыть", accelerator="Ctrl+E", command=self.exit)
        self.menubar.add_cascade(label="Файл", menu=self.filemenu)

        # Меню "Редактировать"
        self.editmenu = Menu(self.menubar, font=("times new roman", 12, "bold"), activebackground="skyblue", tearoff=0)
        self.editmenu.add_command(label="Вырезать", accelerator="Ctrl+X", command=self.cut)
        self.editmenu.add_command(label="Копировать", accelerator="Ctrl+C", command=self.copy)
        self.editmenu.add_command(label="Вставить", accelerator="Ctrl+V", command=self.paste)
        self.filemenu.add_separator()
        self.editmenu.add_command(label="Вставить изображение", accelerator="Ctrl+I", command=self.insert_image)
        self.editmenu.add_command(label="Вставить видео", accelerator="Ctrl+V", command=self.insert_video_link)
        self.editmenu.add_separator()
        self.editmenu.add_command(label="Отменить", accelerator="Ctrl+Z", command=self.undo)
        self.menubar.add_cascade(label="Редактировать", menu=self.editmenu)
        self.editmenu.add_separator()
        self.editmenu.add_command(label="Найти и заменить", accelerator="Ctrl+F", command=self.find_replace)

        #Меню "Вид"
        self.viewmenu = Menu(self.menubar, font=("times new roman", 12, "bold"), activebackground="skyblue", tearoff=0)
        self.viewmenu.add_command(label="Сменить тему", command=self.toggle_theme)
        self.menubar.add_cascade(label="Вид", menu=self.viewmenu)
        self.viewmenu.add_command(label="Увеличить шрифт", accelerator="Ctrl++", command=self.increase_font)
        self.viewmenu.add_command(label="Уменьшить шрифт", accelerator="Ctrl+-", command=self.decrease_font)
        # Создаём меню "Дополнительно", если его ещё нет
        self.extramenu = Menu(self.menubar, font=("times new roman", 12, "bold"), activebackground="skyblue", tearoff=0)
        self.extramenu.add_command(label="Отправить по email", command=self.send_by_email)
        self.menubar.add_cascade(label="Дополнительно", menu=self.extramenu)
        # Меню "Справка"
        self.helpmenu = Menu(self.menubar, font=("times new roman", 12, "bold"), activebackground="skyblue", tearoff=0)
        self.helpmenu.add_command(label="О программе", command=self.infoabout)
        self.menubar.add_cascade(label="Справка", menu=self.helpmenu)

        # Текстовая область и прокрутка
        scrol_y = Scrollbar(self.root, orient=VERTICAL)
        self.txtarea = Text(
            self.root,
            yscrollcommand=scrol_y.set,
            font=("times new roman", 14),
            state="normal",
            relief=GROOVE,
            undo=True
        )
        # Параметры текста
        self.current_font_size = 14
        self.font_family = "times new roman"

        # Начальная тема
        self.is_dark_mode = False
        self.apply_theme()

        self.txtarea.bind("<Button-1>", self.open_video_from_text)
        scrol_y.pack(side=RIGHT, fill=Y)
        scrol_y.config(command=self.txtarea.yview)
        self.txtarea.pack(fill=BOTH, expand=1)

        # Регистрация сочетаний клавиш
        self.shortcuts()

    # Методы класса

    def send_by_email(self, *args):
        try:
            # Получаем текст
            text_content = self.txtarea.get("1.0", "end-1c").strip()
            if not text_content:
                messagebox.showinfo("Отправка", "Документ пуст. Нечего отправлять.")
                return

            # Экранируем текст для URL
            subject = "Документ из текстового редактора"
            body = urllib.parse.quote(text_content, safe='')

            # Формируем mailto-ссылку
            mailto_url = f"mailto:?subject={urllib.parse.quote(subject)}&body={body}"

            # Открываем в системном почтовом клиенте
            webbrowser.open(mailto_url)
            self.status.set("Открыто окно отправки email")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть почтовый клиент:\n{str(e)}")

    def show_shortcuts_help(self, *args):
        help_text = (
            "Горячие клавиши:\n\n"
            "Ctrl+N — Новый файл\n"
            "Ctrl+O — Открыть файл\n"
            "Ctrl+S — Сохранить\n"
            "Ctrl+A — Сохранить как\n"
            "Ctrl+E — Закрыть\n"
            "Ctrl+X — Вырезать\n"
            "Ctrl+C — Копировать\n"
            "Ctrl+V — Вставить\n"
            "Ctrl+Z — Отменить\n"
            "Ctrl+I — Вставить изображение\n"
            "Ctrl+F — Найти и заменить\n"
            "Ctrl++ — Увеличить шрифт\n"
            "Ctrl+- — Уменьшить шрифт\n"
            "Ctrl+T — Сменить тему (светлая/тёмная)\n"
            "F1 — Эта справка"
        )
        messagebox.showinfo("Горячие клавиши", help_text)

    def find_replace(self, *args):
        # Создаём окно поиска
        find_window = Toplevel(self.root)
        find_window.title("Найти и заменить")
        find_window.geometry("400x150+300+300")
        find_window.transient(self.root)
        find_window.focus_force()

        Label(find_window, text="Найти:", font=("times new roman", 12)).grid(row=0, column=0, padx=10, pady=10,
                                                                             sticky=W)
        find_entry = Entry(find_window, width=30, font=("times new roman", 12))
        find_entry.grid(row=0, column=1, padx=10, pady=10)

        Label(find_window, text="Заменить на:", font=("times new roman", 12)).grid(row=1, column=0, padx=10, pady=10,
                                                                                   sticky=W)
        replace_entry = Entry(find_window, width=30, font=("times new roman", 12))
        replace_entry.grid(row=1, column=1, padx=10, pady=10)

        def find_next():
            word = find_entry.get()
            if word:
                start = self.txtarea.search(word, "insert", END)
                if start:
                    end = f"{start}+{len(word)}c"
                    self.txtarea.tag_add("highlight", start, end)
                    self.txtarea.tag_config("highlight", background="yellow")
                    self.txtarea.mark_set("insert", end)
                    self.txtarea.see(end)
                else:
                    messagebox.showinfo("Поиск", "Слово не найдено")

        def replace():
            word = find_entry.get()
            replace_with = replace_entry.get()
            if word:
                content = self.txtarea.get("1.0", END)
                new_content = content.replace(word, replace_with)
                self.txtarea.delete("1.0", END)
                self.txtarea.insert("1.0", new_content)

        Button(find_window, text="Найти далее", command=find_next, width=12).grid(row=2, column=0, padx=10, pady=10)
        Button(find_window, text="Заменить всё", command=replace, width=12).grid(row=2, column=1, padx=10, pady=10)

    def increase_font(self, *args):
        if self.current_font_size < 50:
            self.current_font_size += 1
            self.update_font()

    def decrease_font(self, *args):
        if self.current_font_size > 6:
            self.current_font_size -= 1
            self.update_font()

    def update_font(self):
        self.txtarea.config(font=(self.font_family, self.current_font_size))
        self.status.set(f"Шрифт: {self.current_font_size}pt")

    def toggle_theme(self):
        self.is_dark_mode = not self.is_dark_mode
        self.apply_theme()

    def apply_theme(self):
        if self.is_dark_mode:
            bg_color = "#1e1e1e"
            fg_color = "#ffffff"
            cursor_color = "#ffffff"
            menu_bg = "#2d2d2d"
            menu_fg = "#ffffff"
        else:
            bg_color = "#ffffff"
            fg_color = "#000000"
            cursor_color = "#000000"
            menu_bg = "skyblue"
            menu_fg = "#000000"

        # Применяем к текстовой области
        self.txtarea.config(
            bg=bg_color,
            fg=fg_color,
            insertbackground=cursor_color,
            font=(self.font_family, self.current_font_size)
        )

        # Обновляем строку состояния и заголовок
        self.titlebar.config(bg=menu_bg, fg=menu_fg)
        self.statusbar.config(bg=menu_bg, fg=menu_fg)

        # Обновляем панель меню (ограниченно, но хотя бы цвет фона)
        self.menubar.config(bg=menu_bg, fg=menu_fg)

    def insert_video_link(self, *args):
        try:
            filepath = filedialog.askopenfilename(
                title="Выберите видео файл",
                filetypes=[
                    ("Видео файлы", "*.mp4 *.avi *.mkv *.mov *.wmv"),
                    ("Все файлы", "*.*")
                ]
            )
            if not filepath:
                return

            # Вставляем путь к файлу как текст
            self.txtarea.insert("insert", f"[Видео: {os.path.basename(filepath)}]\n{filepath}\n")
            self.status.set(f"Добавлена ссылка на видео: {os.path.basename(filepath)}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось добавить видео:\n{str(e)}")

    def open_video_from_text(self, event):
        # Получаем координаты клика
        index = self.txtarea.index(f"@{event.x},{event.y}")
        # Получаем всю строку, на которой был клик
        line_start = self.txtarea.index(f"{index} linestart")
        line_end = self.txtarea.index(f"{index} lineend")
        line_content = self.txtarea.get(line_start, line_end)

        # Проверяем, содержит ли строка метку "[Видео:"
        if "[Видео:" in line_content:
            try:
                # Извлекаем путь к файлу (все, что идет после "]")
                # Предполагаем, что путь находится на следующей строке или сразу после ]
                path_start_index = line_content.find("]") + 1
                video_path = line_content[path_start_index:].strip()

                # Если путь пустой, попробуем взять следующую строку
                if not video_path:
                    next_line_start = self.txtarea.index(f"{line_end} +1c linestart")
                    next_line_end = self.txtarea.index(f"{line_end} +1c lineend")
                    video_path = self.txtarea.get(next_line_start, next_line_end).strip()

                # Проверяем, существует ли файл
                if os.path.exists(video_path):
                    # Запускаем файл с помощью системы
                    os.startfile(video_path)  # Для Windows
                    self.status.set(f"Запущено видео: {os.path.basename(video_path)}")
                else:
                    messagebox.showerror("Ошибка", "Файл видео не найден!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось открыть видео:\n{str(e)}")

    def insert_image(self, *args):
        try:
            filepath = filedialog.askopenfilename(
                title="Выберите изображение",
                filetypes=[
                    ("Изображения", "*.png *.jpg *.jpeg *.gif *.bmp"),
                    ("Все файлы", "*.*")
                ]
            )
            if not filepath:
                return

            # Открываем изображение
            img = Image.open(filepath)
            # Масштабируем, чтобы не было огромного изображения
            max_size = 300
            img.thumbnail((max_size, max_size), Image.LANCZOS)

            # Конвертируем в PhotoImage для Tkinter
            photo = ImageTk.PhotoImage(img)

            # Сохраняем ссылку на изображение, чтобы оно не удалилось сборщиком мусора
            self.current_image = photo

            # Вставляем изображение в текущую позицию курсора
            self.txtarea.image_create("insert", image=photo)
            self.status.set(f"Вставлено изображение: {os.path.basename(filepath)}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось вставить изображение:\n{str(e)}")

    def check_spelling(self):
        try:
            import language_tool_python
            tool = language_tool_python.LanguageTool('ru')
            text = self.txtarea.get("1.0", END)
            matches = tool.check(text)
            corrected = language_tool_python.utils.correct(text, matches)

            if messagebox.askyesno("Орфография", "Найдены ошибки. Исправить автоматически?"):
                self.txtarea.delete("1.0", END)
                self.txtarea.insert("1.0", corrected)
            tool.close()
        except Exception as e:
            messagebox.showerror("Ошибка",f"Не удалось проверить текст \n{str(e)}")
    def settitle(self):
        if self.filename:
            self.title.set(self.filename)
        else:
            self.title.set("Без названия")

    def newfile(self, *args):
        self.txtarea.delete("1.0", END)
        self.filename = None
        self.settitle()
        self.status.set("Создан новый файл")

    def openfile(self, *args):
        try:
            self.filename = filedialog.askopenfilename(
                title="Выбрать файл",
                filetypes=(
                    ("Все файлы", "*.*"),
                    ("Текстовый файл", "*.txt"),
                    ("Файлы Python", "*.py")
                )
            )
            if self.filename:
                with open(self.filename, "r", encoding="utf-8") as infile:
                    self.txtarea.delete("1.0", END)
                    self.txtarea.insert(END, infile.read())
                self.settitle()
                self.status.set("Успешно открыт")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def savefile(self, *args):
        try:
            if self.filename:
                data = self.txtarea.get("1.0", END)
                with open(self.filename, "w", encoding="utf-8") as outfile:
                    outfile.write(data)
                self.settitle()
                self.status.set("Успешно сохранён")
            else:
                self.saveasfile()
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def saveasfile(self, *args):
        try:
            # Определяем поддерживаемые форматы
            filetypes = (
                ("Текстовый файл", "*.txt"),
                ("Файл Python", "*.py"),
                ("HTML-файл", "*.html"),
                ("Markdown-файл", "*.md"),
                ("Все файлы", "*.*")
            )
            # Открываем диалог сохранения с фильтром по форматам
            filepath = filedialog.asksaveasfilename(
                title="Сохранить файл как",
                defaultextension=".txt",
                initialfile="Без названия.txt",
                filetypes=filetypes
            )
            if not filepath:
                return  # Пользователь нажал "Отмена"

            # Получаем содержимое текстовой области
            data = self.txtarea.get("1.0", END).rstrip('\n')  # убираем лишний перенос в конце

            # Записываем файл с указанием кодировки
            with open(filepath, "w", encoding="utf-8") as outfile:
                outfile.write(data)

            # Обновляем внутреннее имя файла и интерфейс
            self.filename = filepath
            self.settitle()
            self.status.set("Успешно сохранён")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def save_as_pdf(self, *args):
        try:
            filepath = filedialog.asksaveasfilename(
                title="Сохранить как PDF (СТО САФУ)",
                defaultextension=".pdf",
                initialfile="Документ_СТО.pdf",
                filetypes=[("PDF файлы", "*.pdf")]
            )
            if not filepath:
                return

            from fpdf import FPDF
            from fpdf.enums import Align

            text = self.txtarea.get("1.0", "end-1c")

            # Создаём PDF
            pdf = FPDF(unit="mm", format="A4")
            pdf.add_page()

            # === ПОЛЯ (СТО: лево=30мм, право=15мм, верх/низ=20мм) ===
            pdf.set_margins(left=30, top=20, right=15)
            pdf.set_auto_page_break(auto=True, margin=20)

            # === ШРИФТ ===
            # Используем встроенный Times (без кириллицы) или добавим DejaVu
            try:
                pdf.add_font("Times", "", "Times.ttf", uni=True)
                pdf.set_font("Times", size=14)
            except:
                # Резерв: DejaVu Sans Condensed (поддерживает кириллицу)
                try:
                    pdf.add_font("DejaVu", "", "DejaVuSansCondensed.ttf", uni=True)
                    pdf.set_font("DejaVu", size=14)
                except:
                    pdf.set_font("Arial", size=14)  # fallback

            # === ОБРАБОТКА ТЕКСТА ===
            lines = text.split('\n')
            first_page = True

            for i, line in enumerate(lines):
                # Первая страница — без номера (если нужно по СТО)
                if i == 0:
                    pdf.set_top_margin(20)  # сбрасываем, если был колонтитул

                if line.strip() == "":
                    pdf.ln(7)  # пустая строка
                    continue

                # === КРАСНАЯ СТРОКА (1.25 см = 12.5 мм) ===
                if not line.startswith(" ") and i > 0:  # эвристика: начало абзаца
                    pdf.write(7, " " * 12)  # приблизительно 1.25 см отступ
                    pdf.set_x(30 + 12.5)  # 30мм левое поле + 12.5мм отступ

                # === ВЫРАВНИВАНИЕ ПО ШИРИНЕ ===
                # FPDF не поддерживает justify, но можно эмулировать через multi_cell
                pdf.multi_cell(
                    w=0,
                    h=7 * 1.5,  # межстрочный интервал 1.5 → высота строки = 7 * 1.5 ≈ 10.5
                    txt=line,
                    align=Align.L,  # justify не поддерживается, но можно использовать плагин
                    new_x="LEFT",
                    new_y="NEXT"
                )

            # === НУМЕРАЦИЯ СТРАНИЦ (СТО: внизу по центру, начиная со 2-й страницы) ===
            alias_nb_pages = pdf.alias_nb_pages()
            pdf.set_y(-15)  # 15 мм от низа
            pdf.set_font("", "I", 12)
            pdf.cell(0, 10, f"Страница {pdf.page_no()} из {{nb}}", align=Align.C)

            pdf.output(filepath)
            self.status.set("PDF сохранён по СТО САФУ")
        except Exception as e:
            messagebox.showerror("Ошибка PDF", str(e))

    def save_as_excel(self, *args):
        try:
            filepath = filedialog.asksaveasfilename(
                title="Сохранить как Excel",
                defaultextension=".xlsx",
                initialfile="Документ.xlsx",
                filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")]
            )
            if not filepath:
                return

            text_content = self.txtarea.get("1.0", END).rstrip('\n')
            wb = Workbook()
            ws = wb.active
            ws.title = "Текст"

            # Каждая строка текста — в отдельную строку Excel (столбец A)
            for i, line in enumerate(text_content.split('\n'), start=1):
                ws[f"A{i}"] = line

            wb.save(filepath)
            self.status.set("Успешно сохранено в Excel")
        except Exception as e:
            messagebox.showerror("Ошибка Excel", str(e))

    def exit(self, *args):
        if messagebox.askyesno("ПРЕДУПРЕЖДЕНИЕ", "Ваши несохранённые данные могут быть потеряны!"):
            self.root.destroy()

    def cut(self, *args):
        self.txtarea.event_generate("<<Cut>>")  # Исправлено: английские названия событий

    def copy(self, *args):
        self.txtarea.event_generate("<<Copy>>")

    def paste(self, *args):
        self.txtarea.event_generate("<<Paste>>")

    def undo(self, *args):
        try:
            self.txtarea.edit_undo()
            self.status.set("Действие отменено")
        except:
            self.status.set("Невозможно отменить")

    def infoabout(self):
        messagebox.showinfo("О текстовом редакторе", "\nПростой текстовый редактор"
                            "\nСоздан с использованием Python."
                            "\n Создатель: Виноградова Т.Д."
                            "\n Год создания: 2023-2026"
                            "\n По техническим причинам обращаться к Администратору")

    def shortcuts(self):
        self.txtarea.bind("<Control-n>", self.newfile)
        self.txtarea.bind("<Control-o>", self.openfile)
        self.txtarea.bind("<Control-s>", self.savefile)
        self.txtarea.bind("<Control-a>", self.saveasfile)
        self.txtarea.bind("<Control-e>", self.exit)
        self.txtarea.bind("<Control-x>", self.cut)
        self.txtarea.bind("<Control-c>", self.copy)
        self.txtarea.bind("<Control-v>", self.paste)
        self.txtarea.bind("<Control-z>", self.undo)  # Ctrl+Z — стандарт для отмены
        self.txtarea.bind("<Control-i>", self.insert_image)
        self.txtarea.bind("<Control-q>", self.insert_video_link)
        self.txtarea.bind("<Control-plus>", self.increase_font)
        self.txtarea.bind("<Control-minus>", self.decrease_font)
        self.txtarea.bind("<Control-equal>", self.increase_font)  # На некоторых клавиатурах "=" вместо "+"
        self.txtarea.bind("<Control-f>", self.find_replace)
        self.root.bind("<F1>", self.show_shortcuts_help)  # привязываем к корневому окну

# Запуск приложения
if __name__ == "__main__":
    root = Tk()
    Text_Editor(root)
    root.mainloop()